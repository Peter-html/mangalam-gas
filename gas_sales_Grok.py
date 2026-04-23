import dearpygui.dearpygui as dpg
import openpyxl
import datetime
import os
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, PieChart, Reference
from collections import defaultdict

# ─── Helpers ──────────────────────────────────────────────

def get_total_expense(wb, iso_date):
    total = 0
    for sheet in wb.sheetnames:
        if "-Expense" not in sheet:
            continue
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] == iso_date:
                total += row[3] or 0
    return total

def clean_old_totals(ws):
    for row in range(ws.max_row, 1, -1):
        v = ws.cell(row,1).value
        if v in ("Total Cash","Total Expense","Net Amount"):
            ws.delete_rows(row, 3)
            return

def add_cylinder_header(ws):
    hdr = ["Date","Salesman","Location","Full-Empty",
           "Cylinder","Qty","Rate","Amount",
           "PhonePe","NetBanking","Transaction","Cash"]
    ws.append(hdr)
    for i in range(len(hdr)):
        cell = ws.cell(1,i+1)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFFF00")

def add_customer_header(ws, startrow):
    hdr = ["Serial No.", "Salesperson", "Customer Name",
           "Cylinder Type", "Quantity", "Amount"]
    for idx,header in enumerate(hdr, 1):
        ws.cell(startrow, idx, header)
        ws.cell(startrow, idx).font = Font(bold=True)
        ws.cell(startrow, idx).fill = PatternFill("solid", fgColor="A2C2F7")

def add_totals(ws, total_exp):
    row = ws.max_row
    while row >= 1 and all((ws.cell(row, c).value in [None, ''] for c in range(1, ws.max_column+1))):
        row -= 1
    ws.append([])
    data = list(ws.values)[1:]
    tot_cyl = sum(r[7] or 0 for r in data if len(r) > 7 and isinstance(r[7], (int,float)))
    tot_ph  = sum(r[8] or 0 for r in data if len(r) > 8 and isinstance(r[8], (int,float)))
    tot_tx  = sum(r[10] or 0 for r in data if len(r) > 10 and isinstance(r[10], (int,float)))
    cash    = tot_cyl - tot_ph - tot_tx
    ws.append(["Total Cash", cash])
    ws.append(["Total Expense", total_exp])
    ws.append(["Net Amount", cash - total_exp])
    for cell in ws[ws.max_row-2][:2]:
        cell.fill = PatternFill("solid", fgColor="FFFF00")
        cell.font = Font(bold=True)

def build_expense_charts(ws):
    # Remove all existing charts in this sheet
    ws._charts = []
    # Clear old summary blocks for charts (columns G–K)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=7, max_col=12):
        for cell in row:
            cell.value = None

    # Collect and total up expenses per salesperson and per type
    salespersons, types = defaultdict(float), defaultdict(float)
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[1]: continue
        amt = r[3] if len(r) > 3 and isinstance(r[3], (int,float)) else 0
        person = r[1]
        etype = r[2]
        if amt and person:
            salespersons[person] += amt
        if amt and etype:
            types[etype] += amt
    # Write salesperson summary (G/H columns)
    ws['G1'] = 'Salesperson'
    ws['H1'] = 'Total Expense'
    for idx, (person, total) in enumerate(salespersons.items(), start=2):
        ws[f'G{idx}'] = person
        ws[f'H{idx}'] = total
    # Bar chart: Expenses by Salesperson
    if len(salespersons):
        bar = BarChart()
        bar.title = "Expenses by Salesperson"
        cat = Reference(ws, min_col=7, min_row=2, max_row=1+len(salespersons))
        val = Reference(ws, min_col=8, min_row=2, max_row=1+len(salespersons))
        bar.add_data(val, titles_from_data=False)
        bar.set_categories(cat)
        ws.add_chart(bar, "G6")
    # Write type summary (J/K columns)
    ws['J1'] = 'Expense Type'
    ws['K1'] = 'Total Expense'
    for idx, (etype, total) in enumerate(types.items(), start=2):
        ws[f'J{idx}'] = etype
        ws[f'K{idx}'] = total
    # Pie chart: Expense Split by Type
    if len(types):
        pie = PieChart()
        pie.title = "Expense Split by Type"
        v2 = Reference(ws, min_col=11, min_row=2, max_row=1+len(types))
        cat2 = Reference(ws, min_col=10, min_row=2, max_row=1+len(types))
        pie.add_data(v2, titles_from_data=False)
        pie.set_categories(cat2)
        ws.add_chart(pie, "K6") # Pie is separated right

# ─── Globals ───────────────────────────────────────────────

salesmen = ["Arul Doss","A Sundar","E Doss","E Muthu","Ganesan",
            "kd Mani","Lakshmanan","m Kannan","Pulraj","Suresh",
            "V Muthu","Muthu","Pandram"]
locations = ["Kalakad Inner","Kalakad Outter","Dohinavur","Eruvadi",
             "Nangunri","Vallioor","Tirgurankudi","Maruthukulam",
             "Moolaikaraipetti","Panagudi"]
c_types = ["14.2","19","5.0"]
DAYS    = [f"{d:02d}" for d in range(1,32)]
MONTHS  = [f"{m:02d}" for m in range(1,13)]
YEARS   = [str(y)    for y in range(2025,2031)]
cylinder_rows = []
expense_rows = []
customer_rows_buffer = []

# ─── Customer Sales Dialog ─────────────────────────────

customer_sales_dialog = None
customer_sales_inputs = {}

def open_customer_dialog():
    global customer_sales_dialog, customer_sales_inputs
    if customer_sales_dialog is not None:
        dpg.delete_item(customer_sales_dialog)
    salesperson = dpg.get_value("entry_name")
    customer_sales_inputs = {}
    with dpg.window(
        label="Add Customer Sale",
        modal=True,
        tag="customer_dialog",
        width=410, height=330,
        no_resize=True, no_move=True, no_collapse=True,
    ) as dlg:
        customer_sales_dialog = dlg
        dpg.add_text(f"Salesperson: {salesperson}")
        customer_sales_inputs['salesperson'] = salesperson
        dpg.add_text("Customer Name")
        customer_sales_inputs['customer_name'] = dpg.add_input_text(width=300)
        dpg.add_text("Cylinder Type")
        customer_sales_inputs['cylinder_type'] = dpg.add_combo(
            c_types, width=140, default_value=c_types[0])
        dpg.add_text("Quantity")
        customer_sales_inputs['quantity'] = dpg.add_input_int(width=140, default_value=0)
        dpg.add_text("Amount")
        customer_sales_inputs['amount'] = dpg.add_input_int(width=140, default_value=0)
        dpg.add_button(label="Save", width=120, callback=buffer_customer_sale)
        dpg.add_same_line()
        dpg.add_button(label="Cancel", width=110,
                       callback=lambda s, a: dpg.delete_item(dlg))

def buffer_customer_sale(sender, app_data):
    global customer_sales_dialog, customer_sales_inputs, customer_rows_buffer
    salesperson = dpg.get_value("entry_name")
    customer_name = dpg.get_value(customer_sales_inputs['customer_name']).strip()
    cylinder_type = dpg.get_value(customer_sales_inputs['cylinder_type'])
    quantity = dpg.get_value(customer_sales_inputs['quantity'])
    amount = dpg.get_value(customer_sales_inputs['amount'])
    if not customer_name:
        dpg.set_value("status_text", "❌ Customer name required")
        return
    if not cylinder_type:
        dpg.set_value("status_text", "❌ Select cylinder type")
        return
    if quantity <= 0 or amount <= 0:
        dpg.set_value("status_text", "❌ Enter valid quantity and amount")
        return
    customer_rows_buffer.append({
        'salesperson': salesperson,
        'customer_name': customer_name,
        'cylinder_type': cylinder_type,
        'quantity': quantity,
        'amount': amount,
    })
    dpg.delete_item("customer_dialog")
    dpg.set_value("status_text", f"✅ Customer buffered ({customer_name}) — Click Save Entry to write to Excel.")

# ─── UI Logic ──────────────────────────────────────────────

def rebuild_date(field):
    d = dpg.get_value(f"{field}_day")
    m = dpg.get_value(f"{field}_month")
    y = dpg.get_value(f"{field}_year")
    if d and m and y:
        iso = f"{y}-{m}-{d}"
        formatted = datetime.datetime.strptime(iso,"%Y-%m-%d").strftime("%d/%m/%Y")
        dpg.set_value(field, formatted)
        if field == "entry_date":
            dpg.set_value("expense_date", formatted)
            dpg.set_value("expense_date_day", d)
            dpg.set_value("expense_date_month", m)
            dpg.set_value("expense_date_year", y)

def sync_salesman():
    dpg.set_value("expense_salesman", dpg.get_value("entry_name"))

def add_cylinder(sender, app_data):
    with dpg.group(horizontal=True, parent="cylinder_container", horizontal_spacing=8):
        c   = dpg.add_combo(c_types, width=100)
        q   = dpg.add_input_int(default_value=0, width=150)
        r   = dpg.add_input_int(default_value=0, width=60, step=0)
        rm  = dpg.add_button(label="X", width=20,
                             callback=remove_cylinder_row,
                             user_data=(c,q,r))
        cylinder_rows.append((c,q,r,rm))

def remove_cylinder_row(sender, app_data, user_data):
    c,q,r = user_data
    tup = next(filter(lambda t: t[0]==c, cylinder_rows), None)
    if tup:
        for item in tup:
            dpg.delete_item(item)
        cylinder_rows[:] = [t for t in cylinder_rows if t[0]!=c]

def add_expense(sender, app_data):
    with dpg.group(horizontal=True, parent="expense_container", horizontal_spacing=8):
        t   = dpg.add_input_text(width=150)
        a   = dpg.add_input_int(default_value=0, width=80)
        rm  = dpg.add_button(label="X", width=20,
                             callback=remove_expense_row,
                             user_data=(t,a))
        expense_rows.append((t,a,rm))

def remove_expense_row(sender, app_data, user_data):
    t,a = user_data
    tup = next(filter(lambda t2: t2[0]==t, expense_rows), None)
    if tup:
        for item in tup:
            dpg.delete_item(item)
        expense_rows[:] = [t2 for t2 in expense_rows if t2[0]!=t]

def save_all():
    global customer_rows_buffer
    try:
        d,m,y = dpg.get_value("entry_date").split("/")
        iso    = f"{y}-{m}-{d}"
        datetime.datetime.strptime(iso,"%Y-%m-%d")
    except:
        dpg.set_value("status_text","❌ Invalid Sales date")
        return

    fname = f"gas_sales_{y}_{m}.xlsx"
    wb    = openpyxl.load_workbook(fname) if os.path.exists(fname) else openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    if iso not in wb.sheetnames:
        ws = wb.create_sheet(iso)
        add_cylinder_header(ws)
    else:
        ws = wb[iso]
    clean_old_totals(ws)
    main_end = 2
    for i in range(2, ws.max_row+1):
        if all(ws.cell(i, col).value in ("",None) for col in range(1,ws.max_column+1)):
            main_end = i-1
            break
    if ws.max_row > main_end:
        ws.delete_rows(main_end+1, ws.max_row-main_end)
    loc  = dpg.get_value("entry_location")
    sal  = dpg.get_value("entry_name")
    fe   = dpg.get_value("entry_full_empty") or ""
    ph   = dpg.get_value("entry_phonepe") or 0
    nb   = dpg.get_value("entry_netbanking") or ""
    tx   = dpg.get_value("entry_transaction_amount") or 0
    cyl_data=[]; tot_c=0
    for c,q,r,_ in cylinder_rows:
        typ = dpg.get_value(c) or ""
        qty = dpg.get_value(q) or 0
        rt  = dpg.get_value(r) or 0
        amt = qty*rt
        if typ:
            cyl_data.append((typ,qty,rt,amt)); tot_c+=amt
    cash = tot_c - ph - tx
    base = [iso,sal,loc,fe]
    if cyl_data:
        first,*rest = cyl_data; base += list(first)
    else:
        rest=[]; base+=["",0,0,0]
    base += [ph,nb,tx,cash]
    ws.append(base)
    for row in rest:
        ws.append(["","","",""]+list(row)+["","","",""])
    ws.append([])
    customer_row_start = ws.max_row + 1
    if customer_rows_buffer:
        add_customer_header(ws, customer_row_start)
        sn=1
        for cr in customer_rows_buffer:
            ws.append([sn, cr['salesperson'], cr['customer_name'],
                       cr['cylinder_type'], cr['quantity'], cr['amount']])
            sn+=1
        ws.append([])
    try:
        d2,m2,y2 = dpg.get_value("expense_date").split("/")
        datetime.datetime.strptime(f"{y2}-{m2}-{d2}","%Y-%m-%d")
    except:
        dpg.set_value("status_text","❌ Invalid Expense date")
        return
    exp_data=[]
    for t,a,_ in expense_rows:
        typ = dpg.get_value(t).strip()
        amt = dpg.get_value(a) or 0
        if typ and amt>0:
            exp_data.append((typ,amt))
    mon   = datetime.datetime.strptime(iso,"%Y-%m-%d").strftime("%B")
    esht  = f"{mon}-Expense"
    if esht not in wb.sheetnames:
        ws2 = wb.create_sheet(esht)
        hdr2=["Date","Salesperson","Expense Type","Amount"]
        ws2.append(hdr2)
        for i in range(4):
            c2 = ws2.cell(1,i+1)
            c2.font = Font(bold=True)
            c2.fill = PatternFill("solid", fgColor="FFFF00")
    else:
        ws2 = wb[esht]
    for typ,amt in exp_data:
        ws2.append([iso,sal,typ,amt])
    build_expense_charts(ws2)
    add_totals(ws, get_total_expense(wb,iso))
    wb.save(fname)
    dpg.set_value("status_text", f"✅ Saved {iso}")
    customer_rows_buffer.clear()

# ─── UI and Theming ─────────────────────────────
dpg.create_context()
dpg.create_viewport(title="Gas Sales & Expense Entry System", width=1200, height=800)
dpg.set_viewport_resize_callback(lambda s,a: dpg.configure_item("main", width=a[0], height=a[1]))
with dpg.theme() as th:
    with dpg.theme_component(dpg.mvAll):
        dpg.add_theme_color(dpg.mvThemeCol_WindowBg,   (255,255,255))
        dpg.add_theme_color(dpg.mvThemeCol_ChildBg,    (255,255,255))
        dpg.add_theme_color(dpg.mvThemeCol_FrameBg,    (255,255,255))
        dpg.add_theme_color(dpg.mvThemeCol_FrameBgActive, (255,255,255))
        dpg.add_theme_color(dpg.mvThemeCol_FrameBgHovered, (255,255,255))
        dpg.add_theme_color(dpg.mvThemeCol_PopupBg, (255,255,255))
        dpg.add_theme_color(dpg.mvThemeCol_Border,     (0,0,0))
        dpg.add_theme_style(dpg.mvStyleVar_FrameBorderSize, 1)
        dpg.add_theme_color(dpg.mvThemeCol_Button,     (0,75,160))
        dpg.add_theme_color(dpg.mvThemeCol_ButtonHovered,(30,105,190))
        dpg.add_theme_color(dpg.mvThemeCol_Text,       (0,0,0))
        dpg.add_theme_color(dpg.mvThemeCol_TitleBg,    (255,255,0))
dpg.bind_theme(th)

def build_ui():
    with dpg.window(tag="main", no_title_bar=True, no_move=True, no_resize=True,
                    width=-1, height=-1):
        dpg.add_text("Gas Sales Entry"); dpg.add_separator()
        with dpg.group(horizontal=True, horizontal_spacing=12):
            dpg.add_text("Date (DD/MM/YYYY)")
            dpg.add_input_text(tag="entry_date",
                               default_value=datetime.date.today().strftime("%d/%m/%Y"),
                               width=120, readonly=True)
            dpg.add_combo(DAYS, default_value=DAYS[datetime.date.today().day-1],
                          tag="entry_date_day", width=50,
                          callback=lambda s,a: rebuild_date("entry_date"))
            dpg.add_combo(MONTHS, default_value=MONTHS[datetime.date.today().month-1],
                          tag="entry_date_month", width=50,
                          callback=lambda s,a: rebuild_date("entry_date"))
            dpg.add_combo(YEARS, default_value=YEARS[0],
                          tag="entry_date_year", width=70,
                          callback=lambda s,a: rebuild_date("entry_date"))
            dpg.add_text("Location")
            dpg.add_combo(locations, tag="entry_location",
                          default_value=locations[0], width=150)
            dpg.add_text("Salesperson")
            dpg.add_combo(salesmen, tag="entry_name",
                          default_value=salesmen[0], width=150,
                          callback=lambda s,a: sync_salesman())
        with dpg.group(horizontal=True, horizontal_spacing=12):
            dpg.add_text("Full - Empty")
            dpg.add_input_text(tag="entry_full_empty", width=180)
            dpg.add_text("PhonePe ?")
            dpg.add_input_int(tag="entry_phonepe", default_value=0, width=120)
            dpg.add_text("NetBanking Txn ID")
            dpg.add_input_text(tag="entry_netbanking", width=180)
            dpg.add_text("Transaction Amount ?")
            dpg.add_input_int(tag="entry_transaction_amount", default_value=0, width=150)
        dpg.add_separator()
        dpg.add_text("Cylinder Details")
        dpg.add_child_window(tag="cylinder_container", border=True, height=120)
        add_cylinder(None, None)
        with dpg.group(horizontal=True, horizontal_spacing=10):
            dpg.add_button(label="Add Cylinder", callback=add_cylinder)
            dpg.add_button(label="Add Customer", callback=lambda *_: open_customer_dialog())
        dpg.add_separator()
        dpg.add_text("Expense Entry")
        with dpg.group(horizontal=True, horizontal_spacing=12):
            dpg.add_text("Date (DD/MM/YYYY)")
            dpg.add_input_text(tag="expense_date",
                               default_value=datetime.date.today().strftime("%d/%m/%Y"),
                               width=120, readonly=True)
            dpg.add_combo(DAYS, default_value=DAYS[datetime.date.today().day-1],
                          tag="expense_date_day", width=50,
                          callback=lambda s,a: rebuild_date("expense_date"))
            dpg.add_combo(MONTHS, default_value=MONTHS[datetime.date.today().month-1],
                          tag="expense_date_month", width=50,
                          callback=lambda s,a: rebuild_date("expense_date"))
            dpg.add_combo(YEARS, default_value=YEARS[0],
                          tag="expense_date_year", width=70,
                          callback=lambda s,a: rebuild_date("expense_date"))
            dpg.add_text("Salesperson")
            dpg.add_input_text(tag="expense_salesman", width=150, readonly=True)
        dpg.add_child_window(tag="expense_container", border=True, height=120)
        add_expense(None, None)
        with dpg.group(horizontal=True, horizontal_spacing=10):
            dpg.add_button(label="Add Expense", callback=add_expense)
        dpg.add_separator()
        dpg.add_button(label="Save Entry", callback=save_all)
        dpg.add_text("", tag="status_text")

with dpg.handler_registry():
    dpg.add_key_down_handler(callback=lambda s,a: save_all() if (a==dpg.mvKey_S and dpg.is_key_down(dpg.mvKey_Control)) else None)

build_ui()
dpg.setup_dearpygui()
dpg.show_viewport()
dpg.set_primary_window("main", True)
dpg.start_dearpygui()
dpg.destroy_context()
